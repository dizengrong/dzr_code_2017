# -*- coding: utf-8 -*-
import sys, shutil, os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from util_cell import style_range
from util_cell import set_cell_attribute
from util_sheet import set_col_width
from common import *
import calendar
from datetime import datetime
from util_logger import Logger

# 双：red 单：绿 大：黄 小：(112, 48, 161) 

LIAN_KAI_TIMES = 4


def get_conf_cell_color(val):
    if val == u'双':
        return colors.RED
    elif val == u'单':
        return colors.GREEN
    elif val == u'大':
        return colors.YELLOW
    else: # 小
        return '007030a1'


# 查找连开
def find_lian_kai(ws, col_begin, col_end):
    for row0 in xrange(2, ws.max_row):
        row = ws.max_row - row0 + 2
        for col in xrange(col_begin, col_begin + 1):
            val = ws.cell(row, col).value
            row_add = 0
            times = 1
            cell_list = [(row, col)]
            for c in xrange(col + 1, col_end + 1):
                row_add -= 1
                if row + row_add < 2:
                    break
                # Logger.logger.info("row:%s col: %s val:%s val2:%s" % (row + row_add, c, ws.cell(row + row_add, c).value, val))
                if val == ws.cell(row + row_add, c).value:
                    times += 1
                    cell_list.append((row + row_add, c))
                else:
                    break
            if times >= LIAN_KAI_TIMES:
                color = get_conf_cell_color(val)
                # print("times:", times)
                [add_color_to_cell(ws, r, c, color) for r, c in cell_list]


def add_color_to_cell(ws, row, col, color):
    # print("set cell({}, {}) to color:{}".format(row, col, color))
    fill = PatternFill(fill_type = 'solid', fgColor=color)
    set_cell_attribute(ws, row, col, fill = fill)


def copy_from_src_sheet(src_ws, dest_ws):
    for row in xrange(1, ws.max_row + 1):
        for col in xrange(1, ws.max_column + 1):
            dest_ws.cell(row, col).value = src_ws.cell(row, col).value
            # dest_ws.cell(row, col).font = src_ws.cell(row, col).font
            # dest_ws.cell(row, col).alignment = src_ws.cell(row, col).alignment
            # dest_ws.cell(row, col).border = src_ws.cell(row, col).border
            # dest_ws.cell(row, col).fill = src_ws.cell(row, col).fill



if __name__ == '__main__':
    Logger.setup_logging('logger.yaml')
    src_file = "test.xlsm"
    dst_file = "temp.xlsm"
    shutil.copyfile(src_file, dst_file)
    wb = load_workbook(filename=dst_file, data_only=True, keep_vba=True)
    ws = wb[u'斜数据']
    Logger.logger.info(wb.sheetnames)
    for name in wb.sheetnames:
        # del wb[name]
        if name not in [u'斜数据', u'快乐十分开奖', u'广东快乐十分']:
            del wb[name]

    # new_wb = Workbook()
    # new_ws = new_wb.active
    # new_ws.title = u"斜数据"
    # copy_from_src_sheet(ws, new_ws)

    find_lian_kai(ws, 1, 8)
    find_lian_kai(ws, 9, 16)
    find_lian_kai(ws, 17, 24)

    wb.save("temp.xlsm")
    # wb.save("test2.xlsm")
    # os.remove(dst_file)

