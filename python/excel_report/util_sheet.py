# -*- coding: utf-8 -*-
"""
openxyl操作与sheet有关的方法
"""
from openpyxl.utils import get_column_letter


def set_col_width(ws, col, width):
    col_letter = get_column_letter(col)
    ws.column_dimensions[col_letter].width = width


def set_col_auto_size(ws, col):
    """设置列宽为自动适应，这个应该在已写入保存的文件后调用才会有效"""
    col_letter = get_column_letter(col)
    col = ws.column_dimensions[col_letter]
    col.bestFit = True

