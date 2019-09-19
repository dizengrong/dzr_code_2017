# -*- coding: utf-8 -*-
"""
根据月份生成表头，如下面的这个样式
http://www.360doc.com/content/16/0523/16/20772197_561645721.shtml
"""
import sys
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from util_cell import style_range
from util_cell import set_cell_attribute
from util_sheet import set_col_width
from common import *
import calendar
from datetime import datetime


g_side = Side(border_style='thin', color=colors.WHITE)


def add_day_array(ws, year, month):
    border = Border(left=g_side, top=g_side)
    font = Font(color=colors.WHITE)
    for x in xrange(1, calendar.mdays[month] + 1):
        ws.cell(2, 3 + x).value     = x
        ws.cell(2, 3 + x).font      = font
        ws.cell(2, 3 + x).alignment = Alignment(horizontal = 'center', vertical = 'center')
        ws.cell(2, 3 + x).fill      = PatternFill(fill_type = 'solid', bgColor=colors.BLACK)
        ws.cell(2, 3 + x).border    = border
        set_col_width(ws, 3 + x, 4)


def add_week_array(ws, year, month):
    font   = Font(color=colors.WHITE, size = 8)
    font2  = Font(color=colors.RED, size = 8)
    border = Border(left=g_side, top=g_side)
    for x in xrange(1, calendar.mdays[month] + 1):
        week_day = calendar.weekday(year, month, x)
        ws.cell(3, 3 + x).value = WEEK_DAY_DICT[week_day]
        if week_day == 5 or week_day == 6:
            ws.cell(3, 3 + x).font = font2
        else:
            ws.cell(3, 3 + x).font = font
        ws.cell(3, 3 + x).alignment = Alignment(horizontal = 'center', vertical = 'center')
        ws.cell(3, 3 + x).fill      = PatternFill(fill_type = 'solid', bgColor=colors.BLACK)
        ws.cell(3, 3 + x).border    = border


def gen_table_header(ws, year, month):
    days = calendar.mdays[month]
    font      = Font(color=colors.WHITE, bold=True)
    alignment = Alignment(horizontal = 'center', vertical = 'center', wrapText=True)
    fill      = PatternFill(fill_type = 'solid', bgColor=colors.BLACK)
    border    = Border(left=Side(border_style='thin', color=colors.WHITE))

    ws.merge_cells('A1:A3')
    ws.merge_cells('B1:B3')
    ws.merge_cells('C1:C3')
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=4 + days - 1)
    ws['A1'] = u"序号"
    ws['B1'] = u"部门"
    ws['C1'] = u"姓名"
    ws['D1'] = u"%s年%02d月" % (year, month)
    set_col_width(ws, 1, 6)

    set_cell_attribute(ws, 1, 1, font, alignment, fill, border)
    set_cell_attribute(ws, 1, 2, font, alignment, fill, border)
    set_cell_attribute(ws, 1, 3, font, alignment, fill, border)
    set_cell_attribute(ws, 1, 4, font, alignment, fill, border)
    style_range(ws, 'A1:A3', border=border)
    style_range(ws, 'B1:B3', border=border)
    style_range(ws, 'C1:C3', border=border)

    add_day_array(ws, year, month)
    add_week_array(ws, year, month)

    col = 4 + days
    ws.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
    ws.cell(1, col).value = u'应出\n勤数'
    set_col_width(ws, col, 6)
    col_letter = get_column_letter(col)
    style_range(ws, '%s1:%s3' % (col_letter, col_letter), border, fill, font, alignment)

    col = col + 1
    ws.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
    ws.cell(1, col).value = u'迟到\n次数'
    set_col_width(ws, col, 6)
    col_letter = get_column_letter(col)
    style_range(ws, '%s1:%s3' % (col_letter, col_letter), border, fill, font, alignment)

    col = col + 1
    ws.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
    ws.cell(1, col).value = u'请假\n天数'
    set_col_width(ws, col, 6)
    col_letter = get_column_letter(col)
    style_range(ws, '%s1:%s3' % (col_letter, col_letter), border, fill, font, alignment)

    col = col + 1
    ws.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
    ws.cell(1, col).value = u'实际\n出勤'
    set_col_width(ws, col, 6)
    col_letter = get_column_letter(col)
    style_range(ws, '%s1:%s3' % (col_letter, col_letter), border, fill, font, alignment)

    print "gen_table_header succ"


def usage(module):
    print """usage: %s month""" % (module)


if __name__ == '__main__':
    argv_len = len(sys.argv)
    if argv_len < 2:
        usage(sys.argv[0])
        exit(1)

    year = datetime.now().year
    month = int(sys.argv[1])
    wb = Workbook()
    ws = wb.active
    gen_table_header(ws, year, month)
    wb.save(u"%02d月考勤表头.xlsx" % month)


