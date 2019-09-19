# -*- coding: utf-8 -*-
"""
工作日表
"""
import calendar
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from util_cell import style_range
from util_cell import set_cell_attribute
from common import *
from util_sheet import set_col_width


fill_color_dict = {
    1: PatternFill(fill_type = 'solid', fgColor='00E0FFFF'),
    2: PatternFill(fill_type = 'solid', fgColor='00BFEFFF'),
    3: PatternFill(fill_type = 'solid', fgColor='00CAE1FF'),
    4: PatternFill(fill_type = 'solid', fgColor='00B0E2FF'),
    5: PatternFill(fill_type = 'solid', fgColor='0087CEFF'),
    6: PatternFill(fill_type = 'solid', fgColor='0000BFFF'),
    7: PatternFill(fill_type = 'solid', fgColor='0063B8FF'),
    8: PatternFill(fill_type = 'solid', fgColor='001E90FF'),
    9: PatternFill(fill_type = 'solid', fgColor='004876FF'),
    10: PatternFill(fill_type = 'solid', fgColor='00836FFF'),
    11: PatternFill(fill_type = 'solid', fgColor='00FF83FA'),
    12: PatternFill(fill_type = 'solid', fgColor='00FFBBFF'),
}

g_side = Side(border_style='thin', color=colors.WHITE)

m_font      = Font(bold=True, size=14)
m_alignment = Alignment(horizontal = 'center', vertical = 'center')
m_fill      = PatternFill(fill_type = 'solid', fgColor='00dce6f1')
m_border    = Border(left=g_side, top=g_side, bottom=g_side, right=g_side)

d_font      = Font(bold=True)
d_alignment = Alignment(horizontal = 'center', vertical = 'center')
d_fill      = PatternFill(fill_type = 'solid', fgColor=colors.BLACK)
d_border    = Border(left=Side(border_style='thin', color=colors.WHITE))


def add_month_cell(ws, month):
    row1 = (month - 1) * 3 + 1
    row2 = month * 3
    cell_range = 'A%s:A%s' % (row1, row2)
    ws.merge_cells(cell_range)
    ws['A%s' % row1] = u"%02d月" % month
    style_range(ws, cell_range, border=m_border, alignment=m_alignment, fill=fill_color_dict[month], font=m_font)


def gen_days_table(ws, year):
    for month in xrange(1, 13):
        add_month_cell(ws, month)
        row = (month - 1) * 3 + 1
        for x in xrange(1, calendar.mdays[month] + 1):
            col = 1 + x
            week_day = calendar.weekday(year, month, x)
            col_letter = get_column_letter(col)
            ws['%s%s' % (col_letter, row)].value = x
            set_cell_attribute(ws, row, col, border=m_border, alignment=d_alignment, fill=fill_color_dict[month])

            ws['%s%s' % (col_letter, row + 1)].value = WEEK_DAY_DICT[week_day]
            if week_day == 5 or week_day == 6:
                set_cell_attribute(ws, row + 1, col, border=m_border, alignment=d_alignment, fill=fill_color_dict[month], font=Font(bold=True, color=colors.RED))
            else:
                set_cell_attribute(ws, row + 1, col, border=m_border, alignment=d_alignment, fill=fill_color_dict[month])

            if week_day == 5 or week_day == 6:
                ws['%s%s' % (col_letter, row + 2)].value = u'休'
                set_cell_attribute(ws, row + 2, col, border=m_border, alignment=d_alignment, fill=fill_color_dict[month], font=Font(bold=True, color=colors.RED))
            else:
                ws['%s%s' % (col_letter, row + 2)].value = u'班'
                set_cell_attribute(ws, row + 2, col, border=m_border, alignment=d_alignment, fill=fill_color_dict[month])

    for col in xrange(2, ws.max_column + 1):
        set_col_width(ws, col, 6)


if __name__ == '__main__':
    year = datetime.now().year
    wb = Workbook()
    ws = wb.active
    gen_days_table(ws, year)
    wb.save(u"%s年工作日表.xlsx" % year)
    print 'gen days table succ'

