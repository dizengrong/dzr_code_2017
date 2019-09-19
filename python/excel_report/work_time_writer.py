# -*- coding: utf-8 -*-

"""
考勤表数据转化后写入
"""
import sys
from common import *
from openpyxl import Workbook
import work_time_reader
import table_header
import calendar
from datetime import datetime
from openpyxl.utils import get_column_letter
from util_cell import set_cell_attribute
from openpyxl.styles import Font, Color, Border, Side, Alignment, PatternFill
from openpyxl.styles import colors
from openpyxl.comments import Comment


def get_month_work_days(year, month):
    """
    获取某个月份的应工作天数
    以后可以考虑读取假日表，获取真正的天数
    """
    total = 0
    for day in xrange(1, calendar.mdays[month] + 1):
        week_day = calendar.weekday(year, month, day)
        if not (week_day == 5 or week_day == 6):
            total += 1
    return total


def get_sign_str(sign_dict, date):
    if date not in sign_dict:
        return ''
    return u"打卡时间:\n  " + " - ".join(sign_dict[date])


def get_sign_result(sign_dict, date):
    if date not in sign_dict:
        return u'×'

    time_list = sign_dict[date]
    length = len(time_list)
    if length == 0:
        return u'×'
    elif length == 1:
        return u"漏打"
    else:
        first = time_list[0]
        last = time_list[1]
        if first > '09:00':
            return u'迟'
        elif last < '17:30':
            return u'早退'
        else:
            return u'√'


def write_file(year, month, staff_dict):
    wb = Workbook()
    ws = wb.active
    table_header.gen_table_header(ws, year, month)

    work_days = get_month_work_days(year, month)

    font        = Font(size = 8)
    alignment   = Alignment(horizontal = 'center', vertical = 'center', wrapText=True)
    distributed = Alignment(horizontal = 'distributed', vertical = 'center', wrapText=True)
    fill        = PatternFill(fill_type = 'solid', fgColor='00dce6f1')  # 正常用
    fill2       = PatternFill(fill_type = 'solid', fgColor='0092d050')  # 休息用
    side        = Side(border_style='thin', color=colors.WHITE)
    border      = Border(left=side, top=side, right=side, bottom=side)

    row = 4
    for key in sorted(staff_dict.keys()):
        ws.cell(row, 1).value = key
        ws.cell(row, 3).value = staff_dict[key]['name']
        sign_dict = staff_dict[key]['sign']
        set_cell_attribute(ws, row, 1, alignment = distributed, fill = fill, border = border)
        set_cell_attribute(ws, row, 2, alignment = distributed, fill = fill, border = border)
        set_cell_attribute(ws, row, 3, alignment = distributed, fill = fill, border = border)
        late_times = 0
        leave_times = 0
        for day in xrange(1, calendar.mdays[month] + 1):
            week_day = calendar.weekday(year, month, day)
            col = day + 3
            if week_day == 5 or week_day == 6:
                ws.cell(row, col).value = u'休'
                set_cell_attribute(ws, row, col, font = font, alignment = alignment, fill = fill2, border = border)
            else:
                date = datetime(year, month, day).strftime('%Y-%m-%d')
                ws.cell(row, col).value = get_sign_result(sign_dict, date)
                if ws.cell(row, col).value == u'迟':
                    late_times += 1
                    ws.cell(row, col).comment = Comment(get_sign_str(sign_dict, date), 'system')
                elif ws.cell(row, col).value == u'×':
                    leave_times += 1
                set_cell_attribute(ws, row, col, font = font, alignment = alignment, fill = fill, border = border)

        col = calendar.mdays[month] + 4
        ws.cell(row, col).value = work_days
        set_cell_attribute(ws, row, col, alignment = alignment, fill = fill, border = border)

        col = col + 1
        if late_times > 0:
            ws.cell(row, col).value = late_times
        set_cell_attribute(ws, row, col, alignment = alignment, fill = fill, border = border)

        col = col + 1
        if leave_times > 0:
            ws.cell(row, col).value = leave_times
        set_cell_attribute(ws, row, col, alignment = alignment, fill = fill, border = border)

        col = col + 1
        # ws.cell(row, col).value = work_days - leave_times
        ws.cell(row, col).value = '=%s%s-%s%s' % (get_column_letter(col - 3), row, get_column_letter(col - 1), row)
        set_cell_attribute(ws, row, col, alignment = alignment, fill = fill, border = border)

        row += 1

    wb.save(u"%s年%02d月考勤结果.xlsx" % (year, month))


def usage(module):
    print """usage: %s filename""" % (module)


if __name__ == '__main__':
    argv_len = len(sys.argv)
    if argv_len < 2:
        usage(sys.argv[0])
        exit(1)

    filename = sys.argv[1]
    ret_code, ret_data1, ret_data2 = work_time_reader.read_file(filename)
    if ret_code == RET_CODE_SUCC:
        write_file(ret_data2.year, ret_data2.month, ret_data1)
        print "write data to file succ"
    else:
        print "read excel error happened:%s" % ret_data1

