# -*- coding: utf-8 -*-
"""
原始考勤表数据读取，并转化为需要的格式
"""
import sys
from openpyxl import load_workbook
from datetime import datetime
from common import *


def parse_date(datestring):
    return datetime.strptime(datestring, '%Y-%m-%d')


def read_file(filename):
    wb = load_workbook(filename=filename)
    ws = wb.worksheets[0]
    try:
        print 'total records:%s' % ws.max_row
        staff_dict, work_date = read_file_help(ws)
        print 'read data succ'
        return (RET_CODE_SUCC, staff_dict, work_date)
    except ValueError as e:
        print e.message
        return (RET_CODE_PARSE_EXCEL_VALUE_ERROR, e.message, None)


def read_file_help(ws):
    work_date  = None
    staff_dict = {}
    for row in xrange(2, ws.max_row + 1):
        if work_date is None:
            work_date = parse_date(ws.cell(row, 4).value)
        id        = ws.cell(row, 1).value.strip()
        if not id.isdigit():
            continue
        sign_date = ws.cell(row, 4).value
        sign_time = ws.cell(row, 5).value
        name      = ws.cell(row, 2).value

        if id not in staff_dict:
            staff_dict[id] = {'name': name, 'sign': {}}
        dict = staff_dict[id]['sign']
        # print sign_date
        if sign_date not in dict:
            dict[sign_date] = [sign_time]
        else:
            dict[sign_date].append(sign_time)

    for key in staff_dict.keys():
        sign_dict = staff_dict[key]['sign']
        for key2 in sign_dict.keys():
            time_list = sign_dict[key2]
            if len(time_list) > 2:
                sign_dict[key2] = [time_list[0], time_list[-1]]
    return (staff_dict, work_date)


def usage(module):
    print """usage: %s filename""" % (module)


if __name__ == '__main__':
    argv_len = len(sys.argv)
    if argv_len < 2:
        usage(sys.argv[0])
        exit(1)

    filename = sys.argv[1]
    read_file(filename)


