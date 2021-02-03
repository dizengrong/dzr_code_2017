# -*- coding: utf-8 -*-
import sys
import os
import xlrd
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from tab_lang_ui import Ui_Lang
from xml.dom import minidom
from common import *
import openpyxl
from collections import defaultdict


def make_button(parent, label):
    btn = QPushButton(parent)
    btn.setText(label)
    btn.setFlat(True)
    btn.setStyleSheet("color: rgb(51, 153, 255);")
    btn.setCursor(QCursor(Qt.PointingHandCursor))
    return btn


class TabLang(QtWidgets.QWidget, Ui_Lang):
    def __init__(self, parent, main_window):
        super(TabLang, self).__init__(parent=parent)
        self.main_window = main_window
        self.setupUi(self)
        self.load_lang_conf()
        self.init_table()
        self.init_event()
        self.show_datas()

    def init_table(self):
        # 初始化中文来源表
        header = []
        header.append("Excel文件")
        header.append("Sheet名称")
        for i in range(1,self.max_src_cols - 1):
            header.append("中文列" + str(i))
        self.m_tab_lang.setColumnCount(len(header))

        self.m_tab_lang.setHorizontalHeaderLabels(header)
        self.m_tab_lang.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        self.m_tab_lang.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.m_tab_lang.setSortingEnabled(False)

        # 初始化导出表
        header = []
        header.append("Excel文件")
        header.append("前端")
        self.m_tab_export.setColumnCount(len(header))

        self.m_tab_export.setHorizontalHeaderLabels(header)
        self.m_tab_export.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        self.m_tab_export.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.m_tab_export.setSortingEnabled(False)


    def init_event(self):
        self.m_btn_all.clicked.connect(self.on_fetch_all_lang_src)
        self.m_btn_incr.clicked.connect(self.on_fetch_incr_lang_src)
        self.m_btn_file.clicked.connect(self.on_select_exists_tran_file)

    def load_lang_conf(self):
        # doc = minidom.parse('config/cfg_language.xml')
        doc = minidom.parse(os.path.join(self.main_window.get_config_path(), 'cfg_language.xml'))
        root = doc.documentElement
        self.lang_src = []  # [[excle, sheet, 中文列]]
        self.exports  = []  # [excle]
        self.max_src_cols = 0
        for node in get_xmlnode(root, 'file'):
            excle_file = get_attrvalue(node, 'excle_file')
            if get_attrvalue(node, 'type') == "lang_src":
                for node2 in get_xmlnode(node, 'dict'):
                    temp = [excle_file, get_attrvalue(node2, 'sheet')]
                    for col in get_attrvalue(node2, 'cols').split(','):
                        temp.append(col)
                    self.lang_src.append(temp)
                    self.max_src_cols = max(self.max_src_cols, len(temp))
            elif get_attrvalue(node, 'type') == "translation_src":
                self.exports.append(excle_file)

    def show_datas(self):
        self.m_tab_lang.setRowCount(len(self.lang_src) + 1)
        row = 0
        for item in self.lang_src:
            self.m_tab_lang.setItem(row, 0, QTableWidgetItem(item[0]))
            self.m_tab_lang.setItem(row, 1, QTableWidgetItem(item[1]))
            col = 2
            for n in item[2:]:
                self.m_tab_lang.setItem(row, col, QTableWidgetItem(n))
                col += 1
            row += 1

        self.m_tab_export.setRowCount(len(self.exports) + 1)
        row = 0
        for item in self.exports:
            self.m_tab_export.setItem(row, 0, QTableWidgetItem(item))
            btn = make_button(self.m_tab_export, "导出该语言的所有前端配置")
            btn.setProperty("excle", item)
            self.m_tab_export.setCellWidget(row, 1, btn)
            btn.clicked.connect(self.on_export_all_client)
            row += 1

    def on_export_all_client(self):
        sender = self.sender()
        excel = sender.property("excle")
        excle_filename = os.path.join(self.main_window.get_excel_src_path(), excel)
        # 1.获取已经翻译的文本，这些会在生成时过滤掉
        translate_words = {}
        xml_data = xlrd.open_workbook(excle_filename)
        table = xml_data.sheet_by_index(0)
        for i in range(2, table.nrows):
            translate_words[table.cell(i, 0).value] = table.cell(i, 1).value

        translate_cols = {}
        for data in self.lang_src:
            translate_cols[data[0] + "." + data[1] + "." + data[2]] = True
        self.main_window.m_tab_mod_conf.on_export_all_client_help(translate_cols = translate_cols, translate_words = translate_words)

    def on_fetch_all_lang_src(self):
        filename = QFileDialog.getSaveFileName(self, u"保存文件", filter = "Excel files (*.xlsx)")
        if len(filename[0]) == 0:
            return

        # 提取所有需要翻译的文本
        is_ok, words = self.collect_zh_words()
        if not is_ok:
            return
        # 写入文件中
        self.write_need_translate_words(filename[0], words, {})

    def write_need_translate_words(self, save_filename, words, filter_words):
        to_file = openpyxl.Workbook()
        to_file.remove(to_file.active)  # 删除默认就有的sheet
        to_sheet = to_file.create_sheet(u"translation")
        to_sheet.cell(1, 1, "key")
        to_sheet.cell(1, 2, "value")
        to_sheet.cell(2, 1, "中文文本")
        to_sheet.cell(2, 2, "翻译文本")
        to_sheet.column_dimensions["A"].width = 100
        to_sheet.column_dimensions["B"].width = 100

        fill = openpyxl.styles.PatternFill("solid", fgColor="BCEE68")
        to_sheet.cell(1, 1).fill = fill
        to_sheet.cell(1, 2).fill = fill
        to_sheet.cell(2, 1).fill = fill
        to_sheet.cell(2, 2).fill = fill

        row = 3
        for word in words:
            if word not in filter_words:
                to_sheet.cell(row, 1, word)
                row += 1
            
        to_file.save(save_filename)

    def on_fetch_incr_lang_src(self):
        # 提取增量需要翻译的文本
        filename = QFileDialog.getSaveFileName(self, u"保存文件", filter = "Excel files (*.xlsx)")
        if len(filename[0]) == 0:
            return

        exists_file = self.m_edit_file.text()
        if len(exists_file) == 0:
            msg = u"请先选择对比文件！"
            msg_box = QMessageBox(QMessageBox.Critical, u"错误", msg, parent=self)
            msg_box.exec_()
            return

        # 1.获取已经翻译的文本，这些会在生成时过滤掉
        filter_words = {}
        xml_data = xlrd.open_workbook(exists_file)
        table = xml_data.sheet_by_index(0)
        for i in range(2, table.nrows):
            filter_words[table.cell(i, 0).value] = True

        # 2.提取所有需要翻译的文本，做差异输出
        is_ok, words = self.collect_zh_words()
        if not is_ok:
            return

        # 3.写入文件
        self.write_need_translate_words(filename[0], words, filter_words)

    def on_select_exists_tran_file(self):
        filename = QFileDialog.getOpenFileName(self, u"选择文件", filter = "Excel files (*.xlsx)")
        if len(filename[0]) == 0:
            return
        self.m_edit_file.setText(filename[0])

    def collect_zh_words(self):
        # 获取每个excel配置文件的sheet从哪一行读取数据的配置
        sheet_begin_row_dict = defaultdict(int)  # {file-sheet:begin_row}
        for x in self.main_window.m_tab_mod_conf.export_files.values():
            for sheet in x['datas']:
                sheet_begin_row_dict[x['excle_file'] + '-' + sheet['sheet']] = sheet['begin_row']

        # print(sheet_begin_row_dict)
        words = {}
        for data in self.lang_src:
            excle_filename = os.path.join(self.main_window.get_excel_src_path(), data[0])
            xml_data = xlrd.open_workbook(excle_filename)
            table = xml_data.sheet_by_name(data[1])
            # 根据列名获取中文列对应的col索引
            zh_col = data[2]
            col = 0
            find = False
            for index in range(0, table.ncols):
                if table.cell(0, index).value == zh_col:
                    col = index
                    find = True
                    break
            if not find:
                msg = u"文件【{0}】所在sheet:{1}中找不到列:{2}".format(data[0], data[1], data[2])
                msg_box = QMessageBox(QMessageBox.Critical, u"错误", msg, parent=self)
                msg_box.exec_()
                return False, None

            begin_row = sheet_begin_row_dict[data[0] + '-' + data[1]]
            for i in range(begin_row, table.nrows):
                # 如果第i行的第一列所在的单元格没有数据，则认为是空的，跳过该行
                if str(table.cell(i, 0).value).strip() == '':
                    continue
                words[table.cell(i, col).value] = True
        return True, words
