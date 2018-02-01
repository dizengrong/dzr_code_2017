# -*- coding: utf-8 -*-

import os
import wx
import base_main_gui
import util_excel
import xlrd
import logging
import openpyxl
import traceback
import shutil
import time
import wx.lib.dialogs

# Implementing BaseMainFrame


def show_tips(tips):
    wx.MessageBox(tips, caption=u'提示', style=wx.OK | wx.CENTRE)


def show_error_msg(parent, msg):
    msg_dlg = wx.lib.dialogs.ScrolledMessageDialog(parent, msg, u"错误信息")
    msg_dlg.ShowModal()
    msg_dlg.Destroy()


class MainGui(base_main_gui.BaseMainFrame):
    def __init__(self, parent):
        base_main_gui.BaseMainFrame.__init__(self, parent)
        displaySize = wx.DisplaySize()
        frame_size = self.GetSize()
        self.SetSize(frame_size.GetWidth() + 17, displaySize[1] * 3 / 4)
        self.Centre(wx.BOTH)

        self.reset_data()

    def reset_data(self):
        self.src_files = []
        self.export_sheets = []
        self.max_unicode_cols = 0

    # Handlers for BaseMainFrame events.
    def on_import_src(self, event):
        path = self.m_dir_src_import.GetPath()
        if path == '':
            return
        self.reset_data()
        for f in os.listdir(path):
            if util_excel.is_excel_file(f):
                self.src_files.append(os.path.join(path, f))

        self.on_import_src_help()

    def on_import_src_help(self):
        for f in self.src_files:
            try:
                workbook = xlrd.open_workbook(f)
                for sheet in workbook.sheets():
                    if sheet.name.startswith('draft_'):  # 草稿分页，忽略
                        continue
                    unicode_cols = []
                    for col in xrange(0, sheet.ncols):
                        if util_excel.is_cell_unicode(sheet, 1, col):
                            # logging.debug("sheet %s unicode column:%s" % (sheet.name, sheet.cell(0, col).value))
                            unicode_cols.append((col, sheet.cell(0, col).value))
                    if len(unicode_cols) > 0:
                        self.export_sheets.append((f, sheet.name, unicode_cols))
                        self.max_unicode_cols = max(self.max_unicode_cols, len(unicode_cols))
            except Exception:
                msg = u"导入：%s时失败，可能是该文件中包含不是正常数据的sheet\n\t错误信息:%s\n" % (f, traceback.format_exc())
                show_error_msg(self, msg)
                return
        self.show_export_sheets()

    def show_export_sheets(self):
        self.m_data_result.DeleteAllItems()
        self.m_data_result.ClearColumns()
        self.m_data_result.AppendTextColumn(u'原始文件', align = wx.ALIGN_CENTER, width = wx.COL_WIDTH_AUTOSIZE)
        self.m_data_result.AppendTextColumn(u'sheet名称', align = wx.ALIGN_CENTER, width = wx.COL_WIDTH_AUTOSIZE)
        for index in xrange(0, self.max_unicode_cols):
            self.m_data_result.AppendTextColumn(u'中文字符列%s' % (index + 1), align = wx.ALIGN_CENTER, width = wx.COL_WIDTH_AUTOSIZE)

        for c in self.m_data_result.Columns:
            c.Sortable    = True
            c.Reorderable = True

        for (f, sheet_name, unicode_cols) in self.export_sheets:
            basename = os.path.basename(f)
            values = [basename, sheet_name]
            for (col, col_name) in unicode_cols:
                values.append(u"第%s列: %s" % (col + 1, col_name))
            if len(values) < len(self.m_data_result.Columns):
                for x in xrange(0, len(self.m_data_result.Columns) - len(values)):
                    values.append('')
            self.m_data_result.AppendItem(values)

    def on_export_src(self, event):
        export_dir = self.m_dir_src_export.GetPath()
        if export_dir == '':
            return
        for (f, sheet_name, unicode_cols) in self.export_sheets:
            try:
                self.on_export_src_help(export_dir, f, sheet_name, unicode_cols)
            except Exception:
                msg = u"导出文件：%s时失败\n\t错误信息:%s\n" % (f, traceback.format_exc())
                show_error_msg(self, msg)
                return
        show_tips(u"导出完毕")

    def on_export_src_help(self, to_dir, f, sheet_name, unicode_cols):
        workbook = xlrd.open_workbook(f)
        from_sheet = workbook.sheet_by_name(sheet_name)
        save_path = self.get_save_path(to_dir, f)
        if os.path.exists(save_path):
            to_file = openpyxl.load_workbook(save_path)
        else:
            to_file = openpyxl.Workbook()
            to_file.remove(to_file.active)  # 删除默认就有的sheet

        # 删除以前旧的sheet
        if sheet_name in to_file.sheetnames:
            to_file.remove(to_file[sheet_name])
        to_sheet = to_file.create_sheet(sheet_name)

        cur_col = 1
        # cur_row = 0
        # 写入数据的行索引
        to_sheet.cell(1, 1, 'row')
        for row in xrange(1, from_sheet.nrows):
            to_sheet.cell(row + 1, 1, row)

        for (col, col_name) in unicode_cols:
            to_sheet.cell(1, cur_col + 1, u"%s:%s" % (col, col_name))
            for row in xrange(1, from_sheet.nrows):
                val = from_sheet.cell(row, col).value
                # logging.debug("%s sheet %s cell row:%s col:%s val:%s" % (os.path.basename(f), from_sheet.name, row, col, val))
                to_sheet.cell(row + 1, cur_col + 1, val)
            cur_col += 1

        to_file.save(save_path)

    def get_save_path(self, export_dir, f):
        basename = os.path.basename(f)
        if basename.endswith('.xlsm'):  # 将.xlsm文件改为.xlsx 否则文件无法打开
            name, _ = os.path.splitext(basename)
            basename = name + '.xlsx'
        save_path = os.path.join(export_dir, basename)
        return save_path

    def on_export_src_increase(self, event):
        """导出增量"""
        export_dir = self.m_dir_src_export.GetPath()
        if export_dir == '':
            return
        increase_export_dir = os.path.join(export_dir, 'increase')
        if os.path.exists(increase_export_dir):
            shutil.rmtree(increase_export_dir, onerror = self.remove_file_error_callback)
            time.sleep(1)
        else:
            os.makedirs(increase_export_dir)

        for (f, sheet_name, unicode_cols) in self.export_sheets:
            try:
                self.on_export_src_increase_help(export_dir, increase_export_dir, f, sheet_name, unicode_cols)
            except Exception:
                msg = u"导出文件：%s时失败\n\t错误信息:%s\n" % (f, traceback.format_exc())
                show_error_msg(self, msg)
                return
        show_tips(u"导出完毕")

    def on_export_src_increase_help(self, export_dir, increase_export_dir, f, sheet_name, unicode_cols):
        workbook = xlrd.open_workbook(f)
        from_sheet = workbook.sheet_by_name(sheet_name)

        # 先判断有没有已导出的文件，没有则为新的Excel文件，直接导出src
        compare_f = self.get_save_path(export_dir, f)
        if not os.path.exists(compare_f):
            logging.debug("new excel file, export src...")
            self.on_export_src_help(increase_export_dir, f, sheet_name, unicode_cols)
            return
        logging.debug(f)
        logging.debug(compare_f)
        logging.debug(sheet_name)
        compare_workbook = xlrd.open_workbook(compare_f)
        compare_sheet = compare_workbook.sheet_by_name(sheet_name)

        save_path = self.get_save_path(increase_export_dir, f)
        if os.path.exists(save_path):
            to_file = openpyxl.load_workbook(save_path)
        else:
            to_file = openpyxl.Workbook()
            to_file.remove(to_file.active)  # 删除默认就有的sheet

        # 删除以前旧的sheet
        if sheet_name in to_file.sheetnames:
            to_file.remove(to_file[sheet_name])
        to_sheet = to_file.create_sheet(sheet_name)

        has_data_changed = False
        cur_col = 1
        # 写入数据的行索引
        unicode_col_len = len(unicode_cols)
        for (col, col_name) in unicode_cols:
            cur_row = 1
            to_sheet.cell(cur_row, cur_col, 'row')
            to_sheet.cell(cur_row, cur_col + 1, u"%s:%s" % (col, col_name))
            for row in xrange(1, from_sheet.nrows):
                if self.is_cell_changed(unicode_col_len, from_sheet, row, col, compare_sheet, row, cur_col):
                    has_data_changed = True
                    to_sheet.cell(cur_row + 1, cur_col, row)  # 行索引
                    to_sheet.cell(cur_row + 1, cur_col + 1, from_sheet.cell(row, col).value)  # 中文值
                    cur_row += 1
            if has_data_changed:
                cur_col += 2
        if has_data_changed:
            to_file.save(save_path)

    def is_cell_changed(self, unicode_col_len, from_sheet, row1, col1, compare_sheet, row2, col2):
        # 对比发现如果是多出的行或者列都算改变
        # 值不同也算改变
        # if from_sheet.name == u'Improve说明':
        #     logging.debug("row1:%s col1:%s" % (row1, col1))
        #     logging.debug("row2:%s col2:%s" % (row2, col2))
        #     logging.debug("compare_sheet.nrows: %s" % compare_sheet.nrows)
        #     logging.debug("compare_sheet.ncols: %s" % compare_sheet.ncols)
            # logging.debug("frome val:%s compare val:%s" % (from_sheet.cell(row1, col1).value, compare_sheet.cell(row2, col2).value))
        if row1 >= compare_sheet.nrows or col2 >= compare_sheet.ncols:
            logging.debug("%s new cell row1:%s col1:%s col2:%s nrows:%s ncols:%s" % (from_sheet.name, row1, col1, col2, compare_sheet.nrows, compare_sheet.ncols))
            return True
        elif from_sheet.cell(row1, col1).value != compare_sheet.cell(row2, col2).value:
            logging.debug("frome val:%s compare val:%s" % (from_sheet.cell(row1, col1).value, compare_sheet.cell(row2, col2).value))
            return True
        else:
            return False

    def remove_file_error_callback(exe_func, path, exe_info):
        tips = u"删除文件时失败，路径：%s，错误信息：%s\n请确保删除文件时，该文件已关闭！" % (path, exe_info)
        show_tips(tips)

    def on_import_translate(self, event):
        # TODO: Implement on_import_translate
        pass

    def on_write_translate_to_src(self, event):
        # TODO: Implement on_write_translate_to_src
        pass
